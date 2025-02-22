import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from logger_config import logger

# Constants
XL_TABLE = "xl_table"
XL_SHEET = "xl_sheet"

def validate_single_key(py_dict):
    """
    Validates that a dictionary contains only a single key.
    Logs an error and raises ValueError if multiple keys are present.

    Parameters:
        py_dict (dict): Dictionary to validate.

    Returns:
        bool: True if valid, otherwise raises ValueError.
    """
    if len(py_dict.keys()) != 1:
        error_msg = f"❌ Error: Expected a single key, but multiple were found: {list(py_dict.keys())}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    return True


def add_file_to_load_info(file_info, files_to_load):
    """
    Updates the `files_to_load` dictionary with required files and columns.

    Ensures:
    - The dictionary entry exists before updating.
    - Only unique column names are stored.

    Parameters:
        file_info (dict): Dictionary containing file metadata (name, column mappings, types).
        files_to_load (dict): Dictionary tracking which files need to be processed.

    Returns:
        dict: Updated `files_to_load` dictionary.
    """
    logger.debug(f"Running: add_file_to_load_info")

    # Validate that `file_info` contains only one key
    validate_single_key(file_info)

    for file_name, data_info in file_info.items():
        file_extension = os.path.splitext(file_name)[1].lower()  # Normalize file extension

        if file_extension == ".csv":
            logger.debug(f"Loading CSV: {file_name}")
            # Initialize entry if not exists
            files_to_load.setdefault(file_name, {"cols": set()})

            # Extract column names from mappings and types
            column_mapping_keys = set(data_info.get("column_mapping", {}).keys())
            column_types_keys = set(data_info.get("column_types", {}).keys())

            # Update the set with unique column names
            files_to_load[file_name]["cols"].update(column_mapping_keys)
            files_to_load[file_name]["cols"].update(column_types_keys)

        elif file_extension == ".xlsx":
            # Initialize entry if not exists
            files_to_load.setdefault(file_name, {})

            for xl_type, xl_settings in data_info.items():
                if xl_type not in {XL_TABLE, XL_SHEET}:
                    raise ValueError(f"❌ Error: Unsupported `xl_type`: {xl_type}")

                category_key = "xl_tables" if xl_type == XL_TABLE else "xl_sheets"
                files_to_load[file_name].setdefault(category_key, {})

                xl_name = xl_settings.get("name")
                if not xl_name:
                    raise ValueError(f"❌ Error: Missing name for {xl_type}")

                files_to_load[file_name][category_key].setdefault(xl_name, {"cols": set()})

                # Extract column names from mappings and types
                column_mapping_keys = set(xl_settings.get("column_mapping", {}).keys())
                column_types_keys = set(xl_settings.get("column_types", {}).keys())

                # Update the set with unique column names
                files_to_load[file_name][category_key][xl_name]["cols"].update(column_mapping_keys)
                files_to_load[file_name][category_key][xl_name]["cols"].update(column_types_keys)

        else:
            err_msg = f"❌ Error: Unsupported file extension: {file_extension}"
            logger.error(err_msg)
            raise ValueError(err_msg)

    return files_to_load


def load_excel_sheet(file_path, sheet_name, columns_to_load):
    """
    Loads a specific sheet from an Excel file with selected columns.

    Parameters:
        file_path (str): Path to the Excel file.
        sheet_name (str): Name of the sheet to load.
        columns_to_load (list): List of column names to load.

    Returns:
        pd.DataFrame: DataFrame containing the specified columns from the sheet.

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If the sheet or columns are not found.
    """
    logger.debug(f"file_path:{file_path}")
    logger.debug(f"sheet_name:{sheet_name}")
    logger.debug(f"columns_to_load:{columns_to_load}")

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"❌ Error: File '{file_path}' not found.")

    # Load the sheet with only the required columns
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=columns_to_load, engine="openpyxl")
    except ValueError as e:
        raise ValueError(f"❌ Error: {e}")

    return df

from update_xlsx_data import xl_range_details

def extract_table_from_sheet(sheet, table_range):
    """Extracts a table from an Excel sheet into a DataFrame."""
    (
        _,
        _,
        _,
        start_row_number,
        start_col_number,
        _,
        end_row_number,
        end_col_number,
    ) = xl_range_details(table_range)

    data = [
        [cell.value for cell in row]
        for row in sheet.iter_rows(min_row=start_row_number, max_row=end_row_number, min_col=start_col_number, max_col=end_col_number)
    ]

    df = pd.DataFrame(data[1:], columns=data[0])  # First row as headers

    return df


def load_input_data(input_files_folder: str, input_data_dict: dict) -> dict:
    """
    Loads input data from CSV and Excel files based on a given configuration.

    Args:
        input_files_folder (str): The folder containing input files.
        input_data_dict (dict): Dictionary defining the structure and content to be loaded.

    Returns:
        dict: Updated input_data_dict with loaded data.
    
    Raises:
        FileNotFoundError: If the specified file does not exist.
        ValueError: If an unsupported file type is encountered.
        Exception: For any other errors during processing.
    """

    logger.info("-" * 50)
    logger.info("Starting input data loading process...")

    for file_name, data_config in input_data_dict.items():
        file_path = Path(input_files_folder) / file_name
        file_extension = file_path.suffix.lower()  # Normalize file extension

        if not file_path.exists():
            logger.error(f"File not found: {file_path}")
            raise FileNotFoundError(f"File not found: {file_path}")

        logger.info(f"Processing file: {file_path}")

        if file_extension == '.csv':
            logger.debug("Loading CSV data.")
            column_names = data_config['cols']
            df = pd.read_csv(file_path, usecols=column_names)
            input_data_dict[file_name]["data"] = df


        elif file_extension == '.xlsx':
            wb = load_workbook(file_path, data_only=False)
            logger.info(f"Opened Excel file: {file_path}")

            for xl_type, xl_config in data_config.items():
                if xl_type == "xl_sheets":
                    for sheet_name, cols_dict in xl_config.items():
                        logger.info(f"Loading sheet: {sheet_name}")
                        try:
                            data_pd = load_excel_sheet(
                                file_path=file_path,
                                sheet_name=sheet_name,
                                columns_to_load=list(cols_dict.get('cols', []))
                            )
                            input_data_dict[file_name][xl_type][sheet_name]["data"] = data_pd
                            logger.info(f"Loaded sheet '{sheet_name}' successfully.")

                        except Exception as e:
                            logger.error(f"Error loading sheet '{sheet_name}': {e}")
                            raise

                elif xl_type == "xl_tables":
                    logger.debug("Processing Excel tables...")
                    for table_name, cols_dict in xl_config.items():
                        logger.debug(f"Searching for table: {table_name}")
                        
                        for sheet in wb.worksheets:
                            if table_name in sheet.tables: # to confirm here # to confirm here # to confirm here # to confirm here
                                table = sheet.tables[table_name]
                                table_range = table.ref  # e.g., "A1:C10"

                                if table_range:
                                    logger.info(f"Extracting table: {table_name} from range {table_range}")
                                    df = extract_table_from_sheet(sheet, table_range)

                                    input_data_dict[file_name][xl_type][table_name]["data"] = df
                                    logger.info(f"Table '{table_name}' loaded successfully.")

        else:
            logger.error(f"Unsupported file format: {file_extension}")
            raise ValueError(f"Unsupported file format: {file_extension}")

    logger.info("Input data loading process completed.")
    return input_data_dict


def input_data_loader(input_files_folder, config):
    """
    Loads input data based on the configuration file.

    - Determines which files and columns need to be loaded.
    - Updates `files_to_load` dictionary to track processing requirements.

    Parameters:
        input_files_folder (str): Path to the folder containing input files.
        config (dict): Configuration dictionary specifying data sources.

    Returns:
        dict: Dictionary containing input data.
    """
    for _ in range(2): logger.info("")
    logger.info("-" * 50)
    logger.info("🚀 RUNNING INPUT DATA LOADER")
    logger.info("-" * 50)

    files_to_load = {}

    # Iterate through config to determine required input files
    for output_file, outputs in config.get("output_from_input_dict", {}).items():
        logger.info(f"Identifying data required for output file {output_file}")

        # Process tables
        logger.info("Identifying tables to source from input data")
        for _, file_info in outputs.get("tables", {}).items():
            files_to_load = add_file_to_load_info(file_info, files_to_load)

        # Process sheets
        logger.info("Identifying sheets to source from input data")
        for _, file_info in outputs.get("sheets", {}).items():
            files_to_load = add_file_to_load_info(file_info, files_to_load)

    # Load file data
    input_data_dict = load_input_data(
        input_files_folder=input_files_folder,
        input_data_dict=files_to_load
    )

    logger.info("")

    return input_data_dict
