import os
from openpyxl import load_workbook
import pandas as pd
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter
import logging
from logger_config import logger
import shutil


def check_sheet_table_details(table_details: pd.DataFrame) -> pd.DataFrame:
    """
    Checks for overlapping table regions within a given sheet.

    Args:
        table_details (pd.DataFrame): DataFrame containing table details with columns:
            ["table_name", "start_row_number", "end_row_number"]

    Returns:
        ValueError: If overlapping tables are found.
    """
    try:
        # Ensure required columns exist
        required_columns = {"table_name", "start_row_number", "end_row_number"}
        missing_columns = required_columns - set(table_details.columns)
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")

        overlapping_tables = []

        for _, table in table_details.iterrows():
            table_name = table["table_name"]
            row_start = table["start_row_number"]
            row_end = table["end_row_number"]

            logger.debug(f"🔍 Checking Table: {table_name} (Rows {row_start} - {row_end})")

            # Find potential overlapping tables
            overlaps = table_details[
                (table_details["table_name"] != table_name) &  # Exclude self
                (
                    (table_details["start_row_number"].between(row_start, row_end)) |  # Overlaps at the start
                    (table_details["end_row_number"].between(row_start, row_end)) |  # Overlaps at the end
                    ((table_details["start_row_number"] <= row_start) & (table_details["end_row_number"] >= row_end))
                )  # Fully enclosed
            ]

            if not overlaps.empty:
                overlapping_tables.append((table_name, row_start, row_end))
                logger.warning(f"❌ Overlapping Table Found: {table_name} (Rows {row_start}-{row_end})")

        # Return results
        if overlapping_tables:
            overlap_info = "\n".join([f"Table: {name} (Rows {start}-{end})" for name, start, end in overlapping_tables])
            error_message = f"❌ Overlapping tables detected:\n{overlap_info}"
            logger.error(error_message)
            raise ValueError(error_message)

        # If no issues, just return None (does nothing)
        logger.info("✅ No overlapping tables found.")

    except Exception as e:
        logger.critical(f"Unexpected error in check_sheet_table_details: {e}", exc_info=True)
        raise  # Re-raise the exception for visibility

def check_table_details_in_file(table_details: pd.DataFrame):
    """
    Validates table details in a given DataFrame by checking:
    - Table names are unique.
    - Start row and column numbers are less than their corresponding end values.
    - Calls check_sheet_tables for each unique sheet.

    Args:
        table_details (pd.DataFrame): DataFrame containing table details.
    
    Returns:
        pd.DataFrame: Sorted and validated DataFrame.
    """
    try:
        # Check only one file is provided in the table_details file
        if len(table_details["file_path"].unique()) == 1:
            logger.info("✅ PASS: Only one file provided in table_details.")
        else:
            logger.error("❌ FAIL: More than one file referenced in table_details!")

        # Check uniqueness of table names
        if table_details["table_name"].is_unique:
            logger.info("✅ PASS: All table names are unique.")
        else:
            logger.error("❌ FAIL: Duplicate table names found!")
            duplicate_tables = table_details[table_details.duplicated(subset=["table_name"], keep=False)]
            logger.error(f"Duplicate Table Names:\n{duplicate_tables}")

        # Validate start row/col < end row/col
        invalid_rows = table_details[table_details["start_row_number"] >= table_details["end_row_number"]]
        invalid_cols = table_details[table_details["start_col_number"] >= table_details["end_col_number"]]

        if invalid_rows.empty and invalid_cols.empty:
            logger.info("✅ PASS: All start rows/columns are less than end rows/columns.")
        else:
            logger.error("❌ FAIL: Some start rows/columns are not less than end rows/columns.")
            if not invalid_rows.empty:
                logger.error(f"Invalid row ranges:\n{invalid_rows}")
            if not invalid_cols.empty:
                logger.error(f"Invalid column ranges:\n{invalid_cols}")

        # Process each sheet separately
        for sheet_name in table_details["sheet_name"].unique():
            sheet_table_details = table_details[table_details["sheet_name"] == sheet_name]
            logger.info(f"🔍 Processing sheet: {sheet_name} with {len(sheet_table_details)} tables.")
            try:
                check_sheet_table_details(sheet_table_details)  # Ensure this function is defined elsewhere
            except Exception as e:
                logger.error(f"Error processing sheet '{sheet_name}': {e}", exc_info=True)

        return table_details

    except Exception as e:
        logger.critical(f"Unexpected error in check_table_details_in_file: {e}", exc_info=True)
        return None  # Return None if an error occurs

def check_table_details_across_files(table_details):

    for file in table_details["file_path"]:
        table_details_file = table_details[table_details["file_path"] == file]
        check_table_details_in_file(table_details_file)

    return table_details

def table_details_structure(table_details: pd.DataFrame) -> pd.DataFrame:
    """
    Validates and sorts the table_details DataFrame.

    Args:
        table_details (pd.DataFrame): DataFrame containing table details with columns:
            ["file_path", "sheet_name", "table_name", "start_row_number", "end_row_number",
            "start_col_number", "end_col_number"]
    
    Returns:
        pd.DataFrame: Sorted DataFrame if valid, else raises an exception.
    """
    try:
        # Validate required columns exist
        required_columns = {"file_path", "sheet_name", "table_name", "start_row_number", 
                            "end_row_number", "start_col_number", "end_col_number"}
        missing_columns = required_columns - set(table_details.columns)
        
        if missing_columns:
            raise ValueError(f"Missing required columns: {missing_columns}")

        # Sort by file_path, sheet_name, and table_name
        table_details = table_details.sort_values(by=["file_path", "sheet_name", "start_row_number"]).reset_index(drop=True)

        logger.info("✅ Table structure validated and sorted successfully.")
        return table_details

    except Exception as e:
        logger.critical(f"Error in table_details_structure: {e}", exc_info=True)
        raise  # Re-raise the exception for visibility

def get_excel_table_details(file_path):
    """Extracts table information from an Excel file."""
    wb = load_workbook(file_path, data_only=False)  # Load workbook
    table_info = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]  # Get the worksheet

        # Loop through all tables and get their ranges
        for table_name, table_range in ws.tables.items():
            #start_cell_ref, end_cell_ref = table_range.split(":")  # Extract table range
            start_cell_ref, end_cell_ref = ws.tables[table_name].ref.split(":")

            # Extract column letter and row number
            start_col_letter, start_row_number = coordinate_from_string(start_cell_ref)
            start_col_number = column_index_from_string(start_col_letter)
            end_col_letter, end_row_number = coordinate_from_string(end_cell_ref)
            end_col_number = column_index_from_string(end_col_letter)

            # Store values in a list
            table_info.append({
                "file_path": file_path,
                "sheet_name": sheet_name,
                "table_name": table_name,
                "start_row_number": int(start_row_number),
                "end_row_number": int(end_row_number),
                "start_col_number": int(start_col_number),
                "end_col_number": int(end_col_number)
                
            })

    # Convert list to DataFrame
    table_details = pd.DataFrame(table_info)
    print(f"table_details:{table_details}")

    # Validate and sort table details using the separate function
    table_details = table_details_structure(table_details)
    print(f"table_details:{table_details}")

    # Perform checks
    table_details = check_table_details_in_file(table_details)
    print(f"table_details:{table_details}")

    return wb, table_details

def ws_table_to_df(ws, table_name):
    # Get table from the ws
    ws_table = ws.tables[table_name]

    # Get the table range
    table_range = ws_table.ref
    min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(table_range)

    # Extract table data
    data = [[cell.value for cell in row] for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col)]

    # Convert to pandas DataFrame
    df = pd.DataFrame(data[1:], columns=data[0])  # First row is header
    return df


def align_feed_data(table_df: pd.DataFrame, df_data: pd.DataFrame) -> pd.DataFrame:
    """
    Aligns df_data to match table_df's columns:
    - Compares columns in lowercase.
    - Ensures df_data contains all required columns.
    - Raises an error if any columns are missing.
    - Keeps original column names from table_df in the final DataFrame.

    Parameters:
    table_df (pd.DataFrame): The reference DataFrame with the correct structure.
    df_data (pd.DataFrame): The feed DataFrame to be filtered and aligned.

    Returns:
    pd.DataFrame: The aligned DataFrame with the correct columns and order.

    Raises:
    ValueError: If any required column is missing in df_data.
    """

    # Get the required column names from table_df (original case)
    required_columns_original = table_df.columns.tolist()

    # Convert both table_df and df_data column names to lowercase for comparison
    required_columns_lower = [col.lower() for col in required_columns_original]
    df_data_lower = df_data.rename(columns={col: col.lower() for col in df_data.columns})

    # Check if df_data has all required columns
    missing_columns = [col for col in required_columns_lower if col not in df_data_lower.columns]

    if missing_columns:
        raise ValueError(f"Missing required columns in df_data: {missing_columns}")

    # Reorder and select columns, keeping original Excel column names
    aligned_df = df_data_lower[required_columns_lower]
    aligned_df.columns = required_columns_original  # Restore original column names

    return aligned_df

def extract_table_details(sheet_table_details, table_name):
    """
    Extracts table details from a given DataFrame based on the table name.

    Parameters:
    sheet_table_details (pd.DataFrame): DataFrame containing table details.
    table_name (str): Name of the table to extract details for.

    Returns:
    tuple: (table_row_info, table_start_row, table_start_row_data, 
            table_end_row, table_start_col, table_end_col)
    """
    # Filter table details based on the given table name
    table_row_info = sheet_table_details.loc[sheet_table_details["table_name"] == table_name]
    
    # Extract necessary row and column details
    table_start_row = table_row_info["start_row_number"].iloc[0]
    table_start_row_data = table_start_row + 1  # Skip header row
    table_end_row = table_row_info["end_row_number"].iloc[0]
    table_start_col = table_row_info["start_col_number"].iloc[0]
    table_end_col = table_row_info["end_col_number"].iloc[0]

    # Return values as a tuple
    return (
        table_row_info,
        table_start_row,
        table_start_row_data,
        table_end_row,
        table_start_col,
        table_end_col
    )


def remove_data_from_xl_table(
        sheet_table_details,
        table_name,
        ws,
):
    logger.info("Running: remove_data_from_xl_table")
    (
        table_row_info,
        table_start_row,
        table_start_row_data,
        table_end_row,
        table_start_col,
        table_end_col
    ) = extract_table_details(sheet_table_details, table_name)

    # DELETE ALL DATA ROWS (LEAVE HEADER)
    rows_removed = table_end_row - table_start_row - 1 # Leaving one row in the table
    if rows_removed != 0:
        ws.delete_rows(table_start_row_data, rows_removed)
        logger.info(f"Removed {rows_removed} rows from table '{table_name}', starting from row {table_start_row_data} in sheet '{ws.title}'.")

        # Update table references
        sheet_table_details.loc[sheet_table_details["table_name"] == table_name, "end_row_number"] -= (rows_removed)
        sheet_table_details.loc[sheet_table_details["start_row_number"] > table_start_row, "start_row_number"] -= rows_removed
        sheet_table_details.loc[sheet_table_details["start_row_number"] > table_start_row, "end_row_number"] -= rows_removed

        # Refresh underlying table references in Excel
        sheet_table_details_modified = sheet_table_details.loc[sheet_table_details["start_row_number"] >= table_start_row]
        for _, table in sheet_table_details_modified.iterrows():
            tbl_name = table["table_name"]
            row_start = table["start_row_number"]
            row_end = table["end_row_number"]
            col_start = table["start_col_number"]
            col_end = table["end_col_number"]

            tbl_ref = f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_end}"

            if tbl_name in ws.tables:
                ws.tables[tbl_name].ref = tbl_ref
                logger.info(f"Updated table '{tbl_name}' reference to {tbl_ref}.")
            else:
                logger.warning(f"⚠ Table '{tbl_name}' not found in worksheet '{ws}', skipping reference update.")

    # Ensure first column has null values
    for col_idx in range(table_start_col, table_end_col+1, 1):
        cell_ref = f"{get_column_letter(col_idx)}{table_start_row_data}"
        ws[cell_ref] = None 

    return sheet_table_details, ws


def add_data_to_xl_table(
        sheet_table_details,
        table_name,
        ws,
        aligned_df,
):
    (
        table_row_info,
        table_start_row,
        table_start_row_data,
        table_end_row,
        table_start_col,
        table_end_col
    ) = extract_table_details(sheet_table_details, table_name)

    # Add in the additional rows required
    rows_added = len(aligned_df) - 1 # First row should have been left blank
    logger.info(f"Adding {rows_added} rows to table '{table_name}' from row {table_start_row_data + 1} onwards in sheet '{ws.title}'.")
    if rows_added != 0:
        ws.insert_rows(table_start_row_data + 1, amount=rows_added)

        # Update table references
        sheet_table_details.loc[sheet_table_details["table_name"] == table_name, "end_row_number"] += rows_added
        sheet_table_details.loc[sheet_table_details["start_row_number"] > table_start_row, "start_row_number"] += rows_added
        sheet_table_details.loc[sheet_table_details["start_row_number"] > table_start_row, "end_row_number"] += rows_added

        # Refresh underlying table references in Excel
        sheet_table_details_modified = sheet_table_details.loc[sheet_table_details["start_row_number"] >= table_start_row]
        for _, table in sheet_table_details_modified.iterrows():
            tbl_name = table["table_name"]
            row_start = table["start_row_number"]
            row_end = table["end_row_number"]
            col_start = table["start_col_number"]
            col_end = table["end_col_number"]

            tbl_ref = f"{get_column_letter(col_start)}{row_start}:{get_column_letter(col_end)}{row_end}"

            if tbl_name in ws.tables:
                ws.tables[tbl_name].ref = tbl_ref
                logger.info(f"Updated table '{tbl_name}' reference to {tbl_ref}.")
            else:
                logger.warning(f"⚠ Table '{tbl_name}' not found in worksheet '{ws}', skipping reference update.")

    # Refresh table details extracted - updated values
    (
        table_row_info,
        table_start_row,
        table_start_row_data,
        table_end_row,
        table_start_col,
        table_end_col
    ) = extract_table_details(sheet_table_details, table_name)

    # Add in data into the table
    for row_idx, df_row in enumerate(range(table_start_row_data, table_end_row + 1, 1), start=0):  # df row index starts at 0
        for col_idx, df_col in enumerate(range(table_start_col, table_end_col + 1, 1), start=0):  # df col index starts at 0
            # cell_ref = f"{get_column_letter(col_idx + table_start_col)}{row_idx + table_start_row_data}"
            cell_ref = f"{get_column_letter(df_col)}{df_row}"
            ws[cell_ref] = aligned_df.iloc[row_idx, col_idx]  # Correct indexing

    return sheet_table_details, ws


def copy_template_to_output(xlsx_templates_folder, outputs_folder, template_name):
    """
    Copies an Excel template file from the templates folder to the outputs folder.

    Parameters:
        xlsx_templates_folder (str): Path to the folder containing the template.
        outputs_folder (str): Path to the folder where the output file should be saved.
        template_name (str): Name of the template file (e.g., "template.xlsx").

    Returns:
        str: The path to the copied output file.

    Raises:
        FileNotFoundError: If the template file does not exist.
        IOError: If the file cannot be copied.
    """
    template_path = os.path.join(xlsx_templates_folder, template_name)
    output_path = os.path.join(outputs_folder, template_name)

    if not os.path.exists(template_path):
        raise FileNotFoundError(f"❌ Error: Template file '{template_path}' not found.")

    try:
        shutil.copy(template_path, output_path)  # ✅ Copies the file
        logger.info(f"✅ Template copied to: {output_path}")
        return output_path
    except IOError as e:
        raise IOError(f"❌ Error copying file: {e}")


def get_df_data(data_source, input_data_dict):

    print(f"input_data_dict:{input_data_dict}")

    for file_name, data_config in data_source.items():
        file_extension = os.path.splitext(file_name)[1].lower()  # Normalize file extension

        print(f"file_name:{file_name}")

        if file_extension == '.csv':
            logger.info(f"TO DO - import csv data")

        elif file_extension == '.xlsx':
            for xl_type, xl_config in data_config.items():
                print(f"xl_config:{xl_config}")
                if xl_type == "xl_sheet":
                    input_data = input_data_dict[file_name]["xl_sheets"][xl_config['name']]['data']
                    input_data = input_data.rename(columns=xl_config['column_mapping'])
                    final_columns = list(xl_config['column_mapping'].values())
                    input_data = input_data[final_columns]
                    return input_data

                elif xl_type == "xl_table":
                    print(f"input_data_dict:{input_data_dict}")
                    print(f"file_name:{file_name}")
                    print(f"xl_config['name']:{xl_config['name']}")
                    input_data = input_data_dict[file_name]["xl_tables"][xl_config['name']]['data']
                    print(f"input_data:{input_data}")
                    print(f"xl_config['column_mapping']:{xl_config['column_mapping']}")
                    input_data = input_data.rename(columns=xl_config['column_mapping'])
                    print(f"input_data:{input_data}")
                    final_columns = list(xl_config['column_mapping'].values())
                    input_data = input_data[final_columns]
                    return input_data

                else:
                    err_msg = f"❌ Error: Unsupported xl_type: {xl_type}"
                    logger.error(err_msg)
                    raise ValueError(err_msg)

        else:
            raise ValueError(f"❌ Error: Unsupported file extension: {file_extension}")

    data_df = None
    return data_df


def replace_table_data(
        wb,
        table_details,
        table_name,
        input_data,
):
    logger.info("-" * 50)
    logger.info(f"Replacing data in table {table_name}")

    try:
        # Validate table exists
        if table_name not in table_details["table_name"].values:
            error_message = f"❌ Table '{table_name}' not found in file. Available tables: {table_details['table_name'].unique()}"
            logger.error(error_message)
            raise ValueError(error_message)

        # Retrieve the sheet containing the table
        sheet_name = table_details.loc[table_details["table_name"] == table_name, "sheet_name"].iloc[0]
        sheet_table_details = table_details[table_details["sheet_name"] == sheet_name]
        ws = wb[sheet_name]

        # Get the data for the table in the Excel sheet
        table_df = ws_table_to_df(ws, table_name)

        # Ensure the data provided fits into the Excel table data
        aligned_df = align_feed_data(table_df, input_data)

        # Remove the data from the table
        sheet_table_details, ws = remove_data_from_xl_table(sheet_table_details, table_name, ws)

        # Now add in data where needed
        sheet_table_details, ws = add_data_to_xl_table(sheet_table_details, table_name, ws, aligned_df)

        # Confirm ws has been updated
        logger.info(f"✅ Successfully updated table '{table_name}'.")
        logger.info("")

        table_details[table_details["sheet_name"] == sheet_name] = sheet_table_details

        return wb, table_details

    except Exception as e:
        logger.critical(f"Unexpected error in replace_table_data: {e}", exc_info=True)
        raise


def replace_sheet_data(wb, sheet_name, df):
    """
    Replaces all data in the specified worksheet of an openpyxl workbook with new DataFrame data.
    Ensures that the new DataFrame has exactly the same column names as the original sheet.

    Parameters:
        wb (openpyxl.Workbook): The loaded workbook.
        sheet_name (str): The name of the sheet to replace.
        df (pd.DataFrame): The DataFrame with the new data.

    Returns:
        openpyxl.Workbook: The modified workbook.
    """
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"❌ Error: Sheet '{sheet_name}' not found in workbook.")

    ws = wb[sheet_name]  # Get the worksheet

    # 🔹 Step 1: Extract the original column headers from the sheet
    original_columns = [ws.cell(row=1, column=col_idx).value for col_idx in range(1, ws.max_column + 1)]

    if None in original_columns:
        raise ValueError(f"❌ Error: Sheet '{sheet_name}' contains empty column headers.")

    # 🔹 Step 2: Ensure `df` matches the column names & order
    df = df.copy()  # Avoid modifying the original DataFrame

    # Handle missing columns (Fill with NaN)
    for col in original_columns:
        if col not in df.columns:
            raise ValueError(f"❌ Error: A column is missing in the replacement data {col}.")

    # Remove extra columns not in the original sheet
    df = df[original_columns]

    # 🔹 Step 3: Clear all existing data (keep formatting & formulas intact)
    ws.delete_rows(2, ws.max_row)  # Removes data while keeping headers

    # 🔹 Step 4: Write new data (starting from row 2 to keep headers)
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    return wb  # Return the updated workbook


def add_data_to_files(
        output_from_input_dict,
        input_data_dict,
        xlsx_templates_folder,
        outputs_folder,
        report_date
):
    logger.info("")
    logger.info("-" * 50)
    logger.info("Adding data to files")
    logger.info("-" * 50)

    for template_name, type_config in output_from_input_dict.items():
        output_path=copy_template_to_output(xlsx_templates_folder, outputs_folder, template_name)
        logger.info(f"Adding data to {output_path}.")

        # Load workbook once at the start
        wb, table_details = get_excel_table_details(output_path)

        for output_type, input_config in type_config.items():
            if output_type == 'tables':
                for table_name, data_source in input_config.items():
                    # Load workbook and table details
                    logger.info(f"wb:{wb}")
                    logger.info(f"table_details:{table_details}")

                    input_data = get_df_data(data_source, input_data_dict)
                    wb, table_details = replace_table_data(
                        wb=wb,
                        table_details=table_details,
                        table_name=table_name,
                        input_data=input_data,
                    )
                    # Save the workbook
                    wb.save(output_path)

            elif output_type == 'sheets':

                for sheet_name, data_source in input_config.items():
                    input_data = get_df_data(data_source, input_data_dict)
                    wb = replace_sheet_data(
                        wb=wb,
                        sheet_name=sheet_name,
                        df=input_data
                    )

                # Save the workbook
                wb.save(output_path)

            else:
                error_message = f"❌ Output type '{output_type}' not on of hte accepted values."
                logger.error(error_message)
                raise ValueError(error_message)

        logger.info(f"✅ Successfully updated table '{table_name}' in '{output_path}'.")
